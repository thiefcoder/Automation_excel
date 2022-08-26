#package we need to run this program
import openpyxl as xl
from openpyxl.chart import BarChart , Reference


def main():
    while True:
        filename = input("file name: ")
        if filename.find('.xlsx')!=-1:
            process_workbook(filename)
            break
        else:
            print("wrong format :( ",end="")
            print("try again")
def process_workbook(filename):
    wb = xl.load_workbook(filename)#your excel file locaion
    sheet = wb['Sheet1']
    for row in range(2 , sheet.max_row + 1): #number 1 show us name of table we dont need that :)
        cell = sheet.cell(row,3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row,4) #add new column
        corrected_price_cell.value = corrected_price
    #add chart
    values = Reference(sheet,
                       min_row=2,
                       max_row=sheet.max_row,
                       min_col=4,
                       max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart,'e2')
    wb.save('Book2.xlsx') #enter_your_name.xlsx or last file name :)
    print("all of thing done :)")
if __name__ == '__main__':
    main()